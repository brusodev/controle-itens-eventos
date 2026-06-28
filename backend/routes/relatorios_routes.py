"""
Routes para geração de relatórios do sistema
"""
from flask import Blueprint, jsonify, request, send_file
from models import db, OrdemServico, ItemOrdemServico, Item, EstoqueRegional, MovimentacaoEstoque, Categoria
from datetime import datetime
from io import BytesIO
import openpyxl
from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from reportlab.lib.pagesizes import A4, landscape
from reportlab.lib import colors
from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer, PageBreak
from reportlab.lib.units import cm
from reportlab.pdfbase import pdfmetrics
from reportlab.pdfbase.ttfonts import TTFont
from sqlalchemy import func

from routes.auth_routes import login_requerido

relatorios_bp = Blueprint('relatorios', __name__)


@relatorios_bp.route('/api/relatorios/ordens-servico', methods=['GET'])
@login_requerido
def relatorio_ordens_servico():
    """
    Relatório de Ordens de Serviço com filtros opcionais:
    - data_inicio, data_fim
    - regiao (1-6)
    - contratada
    - servico
    - modulo
    """
    try:
        modulo = request.args.get('modulo', 'coffee')
        query = OrdemServico.query.filter_by(modulo=modulo)
        
        # Filtros
        data_inicio = request.args.get('data_inicio')
        data_fim = request.args.get('data_fim')
        regiao = request.args.get('regiao')
        contratada = request.args.get('contratada')
        servico = request.args.get('servico')
        
        if data_inicio:
            query = query.filter(OrdemServico.data_emissao >= datetime.strptime(data_inicio, '%Y-%m-%d'))
        if data_fim:
            # ✅ Ajustar para o final do dia (23:59:59) para incluir registros de hoje
            dt_fim = datetime.strptime(data_fim, '%Y-%m-%d').replace(hour=23, minute=59, second=59, microsecond=999999)
            query = query.filter(OrdemServico.data_emissao <= dt_fim)
        if regiao:
            query = query.filter(OrdemServico.regiao_estoque == int(regiao))
        if contratada:
            query = query.filter(OrdemServico.detentora.ilike(f'%{contratada}%'))
        if servico:
            query = query.filter(OrdemServico.servico.ilike(f'%{servico}%'))
        
        ordens = query.order_by(OrdemServico.data_emissao.desc()).all()
        
        # Estatísticas
        total_os = len(ordens)
        regioes_atendidas = len(set([os.regiao_estoque for os in ordens if os.regiao_estoque]))
        
        # Contagem por serviço
        servicos = {}
        for os in ordens:
            servicos[os.servico] = servicos.get(os.servico, 0) + 1
        
        # Contagem por contratada
        contratadas = {}
        for os in ordens:
            contratadas[os.detentora] = contratadas.get(os.detentora, 0) + 1
        
        return jsonify({
            'success': True,
            'ordens': [os.to_dict() for os in ordens],
            'estatisticas': {
                'total_os': total_os,
                'regioes_atendidas': regioes_atendidas,
                'por_servico': servicos,
                'por_contratada': contratadas
            }
        })
        
    except Exception as e:
        return jsonify({'success': False, 'error': str(e)}), 500


@relatorios_bp.route('/api/relatorios/estoque-posicao', methods=['GET'])
@login_requerido
def relatorio_estoque_posicao():
    """
    Relatório de posição atual do estoque
    Pode filtrar por categoria ou região
    """
    try:
        modulo = request.args.get('modulo', 'coffee')
        categoria_id = request.args.get('categoria_id')
        regiao = request.args.get('regiao')
        
        query = db.session.query(
            Item.id,
            Item.descricao,
            Item.unidade,
            Categoria.nome.label('categoria_nome'),
            Categoria.natureza,
            EstoqueRegional.regiao_numero,
            EstoqueRegional.quantidade_inicial,
            EstoqueRegional.quantidade_gasto
        ).join(Categoria, Item.categoria_id == Categoria.id)\
         .join(EstoqueRegional, Item.id == EstoqueRegional.item_id)\
         .filter(Categoria.modulo == modulo)
        
        if categoria_id:
            query = query.filter(Item.categoria_id == int(categoria_id))
        if regiao:
            query = query.filter(EstoqueRegional.regiao_numero == int(regiao))
        
        resultados = query.all()
        
        # Processar dados
        estoque_dados = []
        total_itens = 0
        total_valor_inicial = 0
        total_valor_gasto = 0
        
        for r in resultados:
            # ✅ Tratamento seguro de valores
            try:
                # Converter valores nulos ou inválidos para 0
                inicial_str = str(r.quantidade_inicial or '0').strip()
                gasto_str = str(r.quantidade_gasto or '0').strip()
                
                # Evitar valores inválidos como '__'
                if not inicial_str or inicial_str == '__' or not inicial_str.replace(',', '').replace('.', '').replace('-', ''):
                    inicial = 0
                else:
                    inicial = float(inicial_str.replace('.', '').replace(',', '.'))
                
                if not gasto_str or gasto_str == '__' or not gasto_str.replace(',', '').replace('.', '').replace('-', ''):
                    gasto = 0
                else:
                    gasto = float(gasto_str.replace('.', '').replace(',', '.'))
            except (ValueError, AttributeError):
                # Se ainda houver erro, usar 0
                inicial = 0
                gasto = 0
            
            disponivel = inicial - gasto
            percentual_uso = (gasto / inicial * 100) if inicial > 0 else 0
            
            estoque_dados.append({
                'item_id': r.id,
                'descricao': r.descricao,
                'unidade': r.unidade,
                'categoria': r.categoria_nome,
                'natureza': r.natureza,
                'regiao': r.regiao_numero,
                'quantidade_inicial': inicial,
                'quantidade_gasto': gasto,
                'quantidade_disponivel': disponivel,
                'percentual_uso': round(percentual_uso, 2)
            })
            
            total_itens += 1
            total_valor_inicial += inicial
            total_valor_gasto += gasto
        
        return jsonify({
            'success': True,
            'estoque': estoque_dados,
            'resumo': {
                'total_itens': total_itens,
                'total_inicial': round(total_valor_inicial, 2),
                'total_gasto': round(total_valor_gasto, 2),
                'total_disponivel': round(total_valor_inicial - total_valor_gasto, 2),
                'percentual_uso_geral': round((total_valor_gasto / total_valor_inicial * 100) if total_valor_inicial > 0 else 0, 2)
            }
        })
        
    except Exception as e:
        return jsonify({'success': False, 'error': str(e)}), 500


@relatorios_bp.route('/api/relatorios/movimentacoes', methods=['GET'])
@login_requerido
def relatorio_movimentacoes():
    """
    Relatório de movimentações de estoque
    Filtros: data_inicio, data_fim, item_id, regiao, tipo
    """
    try:
        modulo = request.args.get('modulo', 'coffee')
        data_inicio = request.args.get('data_inicio')
        data_fim = request.args.get('data_fim')
        item_id = request.args.get('item_id')
        regiao = request.args.get('regiao')
        tipo = request.args.get('tipo')
        
        query = db.session.query(
            MovimentacaoEstoque,
            Item.descricao,
            OrdemServico.numero_os,
            EstoqueRegional.regiao_numero
        ).join(Item, MovimentacaoEstoque.item_id == Item.id)\
         .join(OrdemServico, MovimentacaoEstoque.ordem_servico_id == OrdemServico.id)\
         .join(EstoqueRegional, MovimentacaoEstoque.estoque_regional_id == EstoqueRegional.id)\
         .filter(OrdemServico.modulo == modulo)
        
        if data_inicio:
            query = query.filter(MovimentacaoEstoque.data_movimentacao >= datetime.strptime(data_inicio, '%Y-%m-%d'))
        if data_fim:
            # ✅ Ajustar para o final do dia (23:59:59)
            dt_fim = datetime.strptime(data_fim, '%Y-%m-%d').replace(hour=23, minute=59, second=59, microsecond=999999)
            query = query.filter(MovimentacaoEstoque.data_movimentacao <= dt_fim)
        if item_id:
            query = query.filter(MovimentacaoEstoque.item_id == int(item_id))
        if regiao:
            query = query.filter(EstoqueRegional.regiao_numero == int(regiao))
        if tipo:
            query = query.filter(MovimentacaoEstoque.tipo == tipo.upper())
        
        movimentacoes = query.order_by(MovimentacaoEstoque.data_movimentacao.desc()).all()
        
        resultado = []
        total_saidas = 0
        total_entradas = 0
        
        for mov, desc, num_os, regiao_num in movimentacoes:
            resultado.append({
                'id': mov.id,
                'data': mov.data_movimentacao.strftime('%d/%m/%Y %H:%M'),
                'item_descricao': desc,
                'numero_os': num_os,
                'regiao': regiao_num,
                'quantidade': mov.quantidade,
                'tipo': mov.tipo,
                'observacao': mov.observacao
            })
            
            if mov.tipo == 'SAIDA':
                total_saidas += mov.quantidade
            else:
                total_entradas += mov.quantidade
        
        return jsonify({
            'success': True,
            'movimentacoes': resultado,
            'resumo': {
                'total_movimentacoes': len(resultado),
                'total_saidas': round(total_saidas, 2),
                'total_entradas': round(total_entradas, 2),
                'saldo': round(total_entradas - total_saidas, 2)
            }
        })
        
    except Exception as e:
        return jsonify({'success': False, 'error': str(e)}), 500


@relatorios_bp.route('/api/relatorios/consumo-por-categoria', methods=['GET'])
@login_requerido
def relatorio_consumo_categoria():
    """
    Relatório consolidado de consumo por categoria
    """
    try:
        modulo = request.args.get('modulo', 'coffee')
        data_inicio = request.args.get('data_inicio')
        data_fim = request.args.get('data_fim')
        
        query = db.session.query(
            Categoria.nome.label('categoria'),
            Categoria.natureza,
            Item.descricao,
            Item.unidade,
            func.sum(ItemOrdemServico.quantidade_total).label('total_consumido'),
            func.count(ItemOrdemServico.id).label('vezes_utilizado')
        ).join(Categoria, Item.categoria_id == Categoria.id)\
         .join(ItemOrdemServico, Item.id == ItemOrdemServico.item_id)\
         .join(OrdemServico, ItemOrdemServico.ordem_servico_id == OrdemServico.id)\
         .filter(Categoria.modulo == modulo)
        
        if data_inicio:
            query = query.filter(OrdemServico.data_emissao >= datetime.strptime(data_inicio, '%Y-%m-%d'))
        if data_fim:
            # ✅ Ajustar para o final do dia (23:59:59) para incluir registros de hoje
            dt_fim = datetime.strptime(data_fim, '%Y-%m-%d').replace(hour=23, minute=59, second=59, microsecond=999999)
            query = query.filter(OrdemServico.data_emissao <= dt_fim)
        
        resultados = query.group_by(
            Categoria.nome,
            Categoria.natureza,
            Item.descricao,
            Item.unidade
        ).order_by(Categoria.nome, func.sum(ItemOrdemServico.quantidade_total).desc()).all()
        
        # Agrupar por categoria
        categorias_resumo = {}
        for r in resultados:
            if r.categoria not in categorias_resumo:
                categorias_resumo[r.categoria] = {
                    'categoria': r.categoria,
                    'natureza': r.natureza,
                    'itens': [],
                    'total_itens_diferentes': 0,
                    'total_consumo': 0
                }
            
            categorias_resumo[r.categoria]['itens'].append({
                'descricao': r.descricao,
                'unidade': r.unidade,
                'total_consumido': float(r.total_consumido or 0),
                'vezes_utilizado': r.vezes_utilizado
            })
            categorias_resumo[r.categoria]['total_itens_diferentes'] += 1
            categorias_resumo[r.categoria]['total_consumo'] += float(r.total_consumido or 0)
        
        return jsonify({
            'success': True,
            'categorias': list(categorias_resumo.values())
        })
        
    except Exception as e:
        return jsonify({'success': False, 'error': str(e)}), 500


@relatorios_bp.route('/api/relatorios/itens-mais-utilizados', methods=['GET'])
@login_requerido
def relatorio_itens_mais_utilizados():
    """
    Top 10 (ou N) itens mais utilizados
    """
    try:
        modulo = request.args.get('modulo', 'coffee')
        limite = int(request.args.get('limite', 10))
        data_inicio = request.args.get('data_inicio')
        data_fim = request.args.get('data_fim')
        grupo = request.args.get('grupo')
        ordenar_por = request.args.get('ordenar_por', 'total_consumido')

        query = db.session.query(
            Item.descricao,
            Item.unidade,
            Categoria.nome.label('categoria'),
            func.sum(ItemOrdemServico.quantidade_total).label('total_consumido'),
            func.count(ItemOrdemServico.id).label('vezes_utilizado')
        ).join(ItemOrdemServico).join(Categoria).join(OrdemServico)\
         .filter(Categoria.modulo == modulo)

        if data_inicio:
            query = query.filter(OrdemServico.data_emissao >= datetime.strptime(data_inicio, '%Y-%m-%d'))
        if data_fim:
            dt_fim = datetime.strptime(data_fim, '%Y-%m-%d').replace(hour=23, minute=59, second=59, microsecond=999999)
            query = query.filter(OrdemServico.data_emissao <= dt_fim)
        if grupo:
            query = query.filter(OrdemServico.regiao_estoque == int(grupo))

        ordem = func.count(ItemOrdemServico.id).desc() if ordenar_por == 'vezes_utilizado' \
            else func.sum(ItemOrdemServico.quantidade_total).desc()

        resultados = query.group_by(
            Item.descricao,
            Item.unidade,
            Categoria.nome
        ).order_by(ordem).limit(limite).all()
        
        ranking = []
        for i, r in enumerate(resultados, 1):
            ranking.append({
                'posicao': i,
                'descricao': r.descricao,
                'unidade': r.unidade,
                'categoria': r.categoria,
                'total_consumido': float(r.total_consumido or 0),
                'vezes_utilizado': r.vezes_utilizado
            })
        
        return jsonify({
            'success': True,
            'ranking': ranking
        })
        
    except Exception as e:
        return jsonify({'success': False, 'error': str(e)}), 500


@relatorios_bp.route('/api/relatorios/itens-mais-utilizados/excel', methods=['GET'])
@login_requerido
def exportar_top_itens_excel():
    try:
        modulo     = request.args.get('modulo', 'coffee')
        limite     = int(request.args.get('limite', 10))
        data_inicio = request.args.get('data_inicio')
        data_fim    = request.args.get('data_fim')
        grupo       = request.args.get('grupo')
        ordenar_por = request.args.get('ordenar_por', 'total_consumido')

        query = db.session.query(
            Item.descricao,
            Item.unidade,
            Categoria.nome.label('categoria'),
            func.sum(ItemOrdemServico.quantidade_total).label('total_consumido'),
            func.count(ItemOrdemServico.id).label('vezes_utilizado')
        ).join(ItemOrdemServico).join(Categoria).join(OrdemServico)\
         .filter(Categoria.modulo == modulo)

        if data_inicio:
            query = query.filter(OrdemServico.data_emissao >= datetime.strptime(data_inicio, '%Y-%m-%d'))
        if data_fim:
            dt_fim = datetime.strptime(data_fim, '%Y-%m-%d').replace(hour=23, minute=59, second=59, microsecond=999999)
            query = query.filter(OrdemServico.data_emissao <= dt_fim)
        if grupo:
            query = query.filter(OrdemServico.regiao_estoque == int(grupo))

        ordem = func.count(ItemOrdemServico.id).desc() if ordenar_por == 'vezes_utilizado' \
            else func.sum(ItemOrdemServico.quantidade_total).desc()

        resultados = query.group_by(
            Item.descricao, Item.unidade, Categoria.nome
        ).order_by(ordem).limit(limite).all()

        GRUPOS_NOME = {1: 'Capital / RMSP', 2: 'Interior', 3: 'Litoral'}

        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = 'Itens Mais Utilizados'

        COR_HEADER  = 'FF4F46E5'
        COR_OURO    = 'FFFFD700'
        COR_PRATA   = 'FFB0C4DE'
        COR_BRONZE  = 'FFCD9B6A'
        COR_PAR     = 'FFF5F3FF'
        COR_IMPAR   = 'FFFFFFFF'

        thin   = Side(style='thin', color='FFCCCCCC')
        borda  = Border(left=thin, right=thin, top=thin, bottom=thin)
        NUM_COLS = 6

        # Linha 1 — Título
        ws.merge_cells(f'A1:{get_column_letter(NUM_COLS)}1')
        c = ws['A1']
        c.value = '🏆 Relatório — Itens Mais Utilizados'
        c.font = Font(bold=True, size=14, color='FFFFFFFF')
        c.fill = PatternFill('solid', fgColor=COR_HEADER)
        c.alignment = Alignment(horizontal='center', vertical='center')
        ws.row_dimensions[1].height = 30

        # Linha 2 — Filtros aplicados
        filtros_txt = []
        if data_inicio or data_fim:
            filtros_txt.append(f"Período: {data_inicio or '...'} → {data_fim or '...'}")
        if grupo:
            filtros_txt.append(f"Grupo: {GRUPOS_NOME.get(int(grupo), grupo)}")
        filtros_txt.append(f"Ordenado por: {'Qtd. Consumida' if ordenar_por != 'vezes_utilizado' else 'Vezes Utilizado'}")
        filtros_txt.append(f"Top {limite}")

        ws.merge_cells(f'A2:{get_column_letter(NUM_COLS)}2')
        c2 = ws['A2']
        c2.value = '  |  '.join(filtros_txt)
        c2.font = Font(italic=True, size=10, color='FF6B7280')
        c2.alignment = Alignment(horizontal='center', vertical='center')
        ws.row_dimensions[2].height = 18

        # Linha 3 — Data de geração
        ws.merge_cells(f'A3:{get_column_letter(NUM_COLS)}3')
        c3 = ws['A3']
        c3.value = f'Gerado em: {datetime.now().strftime("%d/%m/%Y às %H:%M")}'
        c3.font = Font(italic=True, size=9, color='FF9CA3AF')
        c3.alignment = Alignment(horizontal='center', vertical='center')
        ws.row_dimensions[3].height = 16

        ws.append([])  # linha em branco

        # Cabeçalhos
        headers = ['Posição', 'Item', 'Categoria', 'Unidade', 'Qtd. Consumida', 'Vezes Utilizado']
        ws.append(headers)
        header_row = ws.max_row
        for col_idx in range(1, NUM_COLS + 1):
            cell = ws.cell(row=header_row, column=col_idx)
            cell.font = Font(bold=True, color='FFFFFFFF', size=11)
            cell.fill = PatternFill('solid', fgColor=COR_HEADER)
            cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
            cell.border = borda
        ws.row_dimensions[header_row].height = 22

        # Dados
        total_consumido_geral = 0.0
        total_vezes_geral = 0
        for i, r in enumerate(resultados, 1):
            qtd = float(r.total_consumido or 0)
            vezes = r.vezes_utilizado
            total_consumido_geral += qtd
            total_vezes_geral += vezes

            if i == 1:
                bg = COR_OURO
            elif i == 2:
                bg = COR_PRATA
            elif i == 3:
                bg = COR_BRONZE
            else:
                bg = COR_PAR if i % 2 == 0 else COR_IMPAR

            ws.append([i, r.descricao, r.categoria, r.unidade, qtd, vezes])
            row_idx = ws.max_row
            row_fill = PatternFill('solid', fgColor=bg)
            for col_idx in range(1, NUM_COLS + 1):
                cell = ws.cell(row=row_idx, column=col_idx)
                cell.fill = row_fill
                cell.border = borda
                cell.alignment = Alignment(vertical='center')
            ws.cell(row=row_idx, column=1).alignment = Alignment(horizontal='center', vertical='center')
            ws.cell(row=row_idx, column=4).alignment = Alignment(horizontal='center', vertical='center')
            ws.cell(row=row_idx, column=5).number_format = '#,##0.00'
            ws.cell(row=row_idx, column=6).alignment = Alignment(horizontal='center', vertical='center')
            if i <= 3:
                ws.cell(row=row_idx, column=2).font = Font(bold=True)

        # Linha de totais
        ws.append([])
        ws.append(['', 'TOTAL GERAL', '', '', total_consumido_geral, total_vezes_geral])
        tot_row = ws.max_row
        for col_idx in range(1, NUM_COLS + 1):
            cell = ws.cell(row=tot_row, column=col_idx)
            cell.fill = PatternFill('solid', fgColor='FFE0E7FF')
            cell.border = borda
            cell.font = Font(bold=True)
        ws.cell(row=tot_row, column=2).alignment = Alignment(horizontal='center')
        ws.cell(row=tot_row, column=5).number_format = '#,##0.00'
        ws.cell(row=tot_row, column=5).alignment = Alignment(horizontal='center')
        ws.cell(row=tot_row, column=6).alignment = Alignment(horizontal='center')

        # Larguras das colunas
        for i, w in enumerate([10, 42, 28, 14, 20, 18], 1):
            ws.column_dimensions[get_column_letter(i)].width = w

        ws.freeze_panes = f'A{header_row + 1}'

        buffer = BytesIO()
        wb.save(buffer)
        buffer.seek(0)

        grupo_sfx = f'_grupo{grupo}' if grupo else ''
        return send_file(
            buffer,
            mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
            as_attachment=True,
            download_name=f'itens_mais_utilizados{grupo_sfx}_{datetime.now().strftime("%Y%m%d_%H%M%S")}.xlsx'
        )
    except Exception as e:
        return jsonify({'success': False, 'error': str(e)}), 500


# ─────────────────────────────────────────────────────
# Helpers para relatório de pagamentos
# ─────────────────────────────────────────────────────

def _parse_vencimento_date(venc_str):
    if not venc_str or not venc_str.strip():
        return None
    for fmt in ('%d/%m/%Y', '%Y-%m-%d', '%d/%m/%y'):
        try:
            return datetime.strptime(venc_str.strip(), fmt).date()
        except ValueError:
            continue
    return None


def _calcular_valor_total_os(os):
    total = 0.0
    for item in os.itens:
        try:
            qtd = float(item.quantidade_total or 0)
            val_str = str(item.valor_unitario or '0').strip()
            if val_str and val_str not in ('0', '__', ''):
                if ',' in val_str:
                    # Formato BR "1.234,56": ponto é milhar, vírgula é decimal
                    val_str = val_str.replace('.', '').replace(',', '.')
                # Sem vírgula (ex: "1234.56"): ponto já é o separador decimal
                val = float(val_str)
            else:
                val = 0.0
            total += qtd * val
        except (ValueError, TypeError):
            pass
    return round(total, 2)


def _get_status_pagamento(os, hoje):
    if os.pagamento_pago:
        return 'pago'
    if not os.pagamento_vencimento or not os.pagamento_vencimento.strip():
        return 'sem_prazo'
    venc = _parse_vencimento_date(os.pagamento_vencimento)
    if venc is None:
        return 'sem_prazo'
    return 'vencido' if venc < hoje else 'pendente'


def _build_pagamento_query(args):
    modulo = args.get('modulo', '')
    empresa = args.get('empresa', '')
    query = OrdemServico.query
    if modulo:
        query = query.filter_by(modulo=modulo)
    if empresa:
        query = query.filter(OrdemServico.detentora.ilike(f'%{empresa}%'))
    return query.order_by(OrdemServico.data_emissao.desc()).all()


def _filtrar_e_montar_pagamentos(ordens, args):
    status_filtro = args.get('status', '')
    data_inicio_str = args.get('data_inicio_vencimento', '')
    data_fim_str = args.get('data_fim_vencimento', '')

    data_inicio = None
    data_fim = None
    if data_inicio_str:
        try:
            data_inicio = datetime.strptime(data_inicio_str, '%Y-%m-%d').date()
        except ValueError:
            data_inicio = _parse_vencimento_date(data_inicio_str)
    if data_fim_str:
        try:
            data_fim = datetime.strptime(data_fim_str, '%Y-%m-%d').date()
        except ValueError:
            data_fim = _parse_vencimento_date(data_fim_str)

    hoje = datetime.now().date()
    resultado = []

    for os in ordens:
        status = _get_status_pagamento(os, hoje)

        if status_filtro and status != status_filtro:
            continue

        venc_date = _parse_vencimento_date(os.pagamento_vencimento)
        if data_inicio and venc_date and venc_date < data_inicio:
            continue
        if data_fim and venc_date and venc_date > data_fim:
            continue
        if (data_inicio or data_fim) and venc_date is None:
            continue

        resultado.append({
            'id': os.id,
            'numeroOS': os.numero_os,
            'empresa': os.detentora or '-',
            'modulo': os.modulo or '-',
            'regiao': os.regiao_estoque or '-',
            'dataEmissao': os.data_emissao.strftime('%d/%m/%Y') if os.data_emissao else '-',
            'vencimento': os.pagamento_vencimento or '-',
            'valorTotal': _calcular_valor_total_os(os),
            'status': status,
        })

    return resultado


@relatorios_bp.route('/api/relatorios/pagamentos', methods=['GET'])
@login_requerido
def relatorio_pagamentos():
    try:
        ordens = _build_pagamento_query(request.args)
        dados = _filtrar_e_montar_pagamentos(ordens, request.args)

        contadores = {'pago': 0, 'pendente': 0, 'vencido': 0, 'sem_prazo': 0}
        valor_total_geral = 0.0
        for d in dados:
            contadores[d['status']] = contadores.get(d['status'], 0) + 1
            valor_total_geral += d['valorTotal']

        return jsonify({
            'success': True,
            'pagamentos': dados,
            'estatisticas': {
                'total': len(dados),
                'pagos': contadores['pago'],
                'pendentes': contadores['pendente'],
                'vencidos': contadores['vencido'],
                'sem_prazo': contadores['sem_prazo'],
                'valor_total': round(valor_total_geral, 2),
            }
        })
    except Exception as e:
        return jsonify({'success': False, 'error': str(e)}), 500


@relatorios_bp.route('/api/relatorios/pagamentos/excel', methods=['GET'])
@login_requerido
def exportar_pagamentos_excel():
    try:
        ordens = _build_pagamento_query(request.args)
        dados = _filtrar_e_montar_pagamentos(ordens, request.args)

        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = 'Controle de Pagamentos'

        COR_HEADER_BG   = 'FF4F46E5'
        COR_PAGO_BG     = 'FFD4EDDA'
        COR_PENDENTE_BG = 'FFFFF3CD'
        COR_VENCIDO_BG  = 'FFF8D7DA'
        COR_SEMPRAZO_BG = 'FFE9ECEF'
        COR_PAGO_FT     = 'FF155724'
        COR_PENDENTE_FT = 'FF856404'
        COR_VENCIDO_FT  = 'FF721C24'
        COR_SEMPRAZO_FT = 'FF495057'

        thin = Side(style='thin', color='FFCCCCCC')
        borda = Border(left=thin, right=thin, top=thin, bottom=thin)

        # Título
        ws.merge_cells('A1:H1')
        c = ws['A1']
        c.value = 'Relatório de Controle de Pagamentos'
        c.font = Font(bold=True, size=14, color='FFFFFFFF')
        c.fill = PatternFill('solid', fgColor=COR_HEADER_BG)
        c.alignment = Alignment(horizontal='center', vertical='center')
        ws.row_dimensions[1].height = 28

        ws.merge_cells('A2:H2')
        c2 = ws['A2']
        c2.value = f'Gerado em: {datetime.now().strftime("%d/%m/%Y às %H:%M")}'
        c2.font = Font(italic=True, size=10, color='FF6B7280')
        c2.alignment = Alignment(horizontal='center')

        # Resumo
        contadores = {'pago': 0, 'pendente': 0, 'vencido': 0, 'sem_prazo': 0}
        valor_total_geral = 0.0
        for d in dados:
            contadores[d['status']] = contadores.get(d['status'], 0) + 1
            valor_total_geral += d['valorTotal']

        ws.merge_cells('A3:H3')
        cr = ws['A3']
        cr.value = (
            f"Total: {len(dados)}  |  Pagos: {contadores['pago']}  |  "
            f"Pendentes: {contadores['pendente']}  |  Vencidos: {contadores['vencido']}  |  "
            f"Sem Prazo: {contadores['sem_prazo']}  |  "
            f"Valor Total: R$ {valor_total_geral:,.2f}"
        )
        cr.font = Font(bold=True, size=10)
        cr.fill = PatternFill('solid', fgColor='FFF3F4F6')
        cr.alignment = Alignment(horizontal='center')
        ws.row_dimensions[3].height = 18

        ws.append([])  # linha em branco

        # Cabeçalhos
        headers = ['Nº O.S.', 'Empresa', 'Módulo', 'Região', 'Data Emissão', 'Vencimento', 'Valor Total (R$)', 'Status']
        ws.append(headers)
        header_row = ws.max_row
        for col_idx in range(1, len(headers) + 1):
            cell = ws.cell(row=header_row, column=col_idx)
            cell.font = Font(bold=True, color='FFFFFFFF', size=11)
            cell.fill = PatternFill('solid', fgColor=COR_HEADER_BG)
            cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
            cell.border = borda
        ws.row_dimensions[header_row].height = 22

        STATUS_LABELS = {
            'pago':      ('Pago',      COR_PAGO_BG,     COR_PAGO_FT),
            'pendente':  ('Pendente',  COR_PENDENTE_BG, COR_PENDENTE_FT),
            'vencido':   ('VENCIDO',   COR_VENCIDO_BG,  COR_VENCIDO_FT),
            'sem_prazo': ('Sem Prazo', COR_SEMPRAZO_BG, COR_SEMPRAZO_FT),
        }

        for d in dados:
            label, bg_color, ft_color = STATUS_LABELS.get(d['status'], ('?', 'FFFFFFFF', 'FF000000'))
            ws.append([
                d['numeroOS'], d['empresa'], d['modulo'], d['regiao'],
                d['dataEmissao'], d['vencimento'], d['valorTotal'], label,
            ])
            row_idx = ws.max_row
            row_fill = PatternFill('solid', fgColor=bg_color)
            for col_idx in range(1, len(headers) + 1):
                cell = ws.cell(row=row_idx, column=col_idx)
                cell.fill = row_fill
                cell.border = borda
                cell.alignment = Alignment(vertical='center')

            ws.cell(row=row_idx, column=7).number_format = 'R$ #,##0.00'
            sc = ws.cell(row=row_idx, column=8)
            sc.font = Font(bold=True, color=ft_color)
            sc.alignment = Alignment(horizontal='center', vertical='center')

        for i, w in enumerate([12, 40, 14, 10, 16, 16, 20, 14], 1):
            ws.column_dimensions[get_column_letter(i)].width = w

        ws.freeze_panes = f'A{header_row + 1}'

        buffer = BytesIO()
        wb.save(buffer)
        buffer.seek(0)

        return send_file(
            buffer,
            mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
            as_attachment=True,
            download_name=f'controle_pagamentos_{datetime.now().strftime("%Y%m%d_%H%M%S")}.xlsx'
        )
    except Exception as e:
        return jsonify({'success': False, 'error': str(e)}), 500


@relatorios_bp.route('/api/relatorios/organizacao/eventos', methods=['GET'])
@login_requerido
def relatorio_organizacao_eventos():
    try:
        grupo      = request.args.get('grupo', '')
        status     = request.args.get('status', '')
        data_inicio = request.args.get('data_inicio', '')
        data_fim    = request.args.get('data_fim', '')
        empresa    = request.args.get('empresa', '')

        query = OrdemServico.query.filter_by(modulo='organizacao')

        if grupo:
            query = query.filter(OrdemServico.regiao_estoque == int(grupo))
        if status:
            query = query.filter(OrdemServico.status == status)
        if data_inicio:
            query = query.filter(OrdemServico.data_emissao >= datetime.strptime(data_inicio, '%Y-%m-%d'))
        if data_fim:
            dt_fim = datetime.strptime(data_fim, '%Y-%m-%d').replace(hour=23, minute=59, second=59)
            query = query.filter(OrdemServico.data_emissao <= dt_fim)
        if empresa:
            query = query.filter(OrdemServico.detentora.ilike(f'%{empresa}%'))

        ordens = query.order_by(OrdemServico.data_emissao.desc()).all()

        GRUPOS = {1: 'Capital/RMSP', 2: 'Interior', 3: 'Litoral'}
        CAT_MAP = {
            'montagem_decoracao':         'Montagem/Decoração',
            'recursos_humanos':           'Recursos Humanos',
            'equipamento_informatica':    'Equipamentos TI',
            'material_grafico_expediente':'Material Gráfico',
        }

        def parse_valor(s):
            try:
                v = str(s or '0').strip()
                if v in ('', '0', '__'):
                    return 0.0
                if ',' in v:
                    # Formato BR "1.234,56": ponto é milhar, vírgula é decimal
                    v = v.replace('.', '').replace(',', '.')
                # Sem vírgula (ex: "1234.56"): ponto já é o separador decimal
                return float(v)
            except (ValueError, TypeError):
                return 0.0

        def custo_categoria(itens, cat_slug):
            return round(sum(
                (float(i.quantidade_total or 0)) * parse_valor(i.valor_unitario)
                for i in itens if (i.categoria or '').lower() == cat_slug
            ), 2)

        eventos = []
        total_pessoas = 0
        custo_total_geral = 0.0

        for os in ordens:
            cat_custos = {slug: custo_categoria(os.itens, slug) for slug in CAT_MAP}
            custo_total = round(sum(cat_custos.values()), 2)
            qtd_pessoas = os.qtd_pessoas_atendidas or 0
            custo_por_pessoa = round(custo_total / qtd_pessoas, 2) if qtd_pessoas > 0 else 0.0

            total_pessoas += qtd_pessoas
            custo_total_geral += custo_total

            eventos.append({
                'id': os.id,
                'numeroOS': os.numero_os,
                'evento': os.evento or '-',
                'dataEvento': os.data or '-',
                'local': os.local or '-',
                'grupo': os.regiao_estoque or '-',
                'grupoNome': GRUPOS.get(os.regiao_estoque, f'Grupo {os.regiao_estoque}'),
                'empresa': os.detentora or '-',
                'status': os.status or 'emitida',
                'dataEmissao': os.data_emissao.strftime('%d/%m/%Y') if os.data_emissao else '-',
                'qtdPessoas': qtd_pessoas,
                'custoMontagem': cat_custos['montagem_decoracao'],
                'custoRH': cat_custos['recursos_humanos'],
                'custoTI': cat_custos['equipamento_informatica'],
                'custoGrafico': cat_custos['material_grafico_expediente'],
                'custoTotal': custo_total,
                'custoPorPessoa': custo_por_pessoa,
                'totalItens': len(os.itens),
            })

        return jsonify({
            'success': True,
            'eventos': eventos,
            'estatisticas': {
                'total_eventos': len(eventos),
                'total_pessoas': total_pessoas,
                'custo_total_geral': round(custo_total_geral, 2),
                'custo_medio_evento': round(custo_total_geral / len(eventos), 2) if eventos else 0.0,
                'custo_medio_pessoa': round(custo_total_geral / total_pessoas, 2) if total_pessoas > 0 else 0.0,
            }
        })
    except Exception as e:
        return jsonify({'success': False, 'error': str(e)}), 500


@relatorios_bp.route('/api/relatorios/organizacao/excel', methods=['GET'])
@login_requerido
def exportar_organizacao_excel():
    try:
        # Reaproveitamos a lógica da rota JSON
        ordens = OrdemServico.query.filter_by(modulo='organizacao')

        grupo       = request.args.get('grupo', '')
        status      = request.args.get('status', '')
        data_inicio = request.args.get('data_inicio', '')
        data_fim    = request.args.get('data_fim', '')
        empresa     = request.args.get('empresa', '')

        if grupo:
            ordens = ordens.filter(OrdemServico.regiao_estoque == int(grupo))
        if status:
            ordens = ordens.filter(OrdemServico.status == status)
        if data_inicio:
            ordens = ordens.filter(OrdemServico.data_emissao >= datetime.strptime(data_inicio, '%Y-%m-%d'))
        if data_fim:
            dt_fim = datetime.strptime(data_fim, '%Y-%m-%d').replace(hour=23, minute=59, second=59)
            ordens = ordens.filter(OrdemServico.data_emissao <= dt_fim)
        if empresa:
            ordens = ordens.filter(OrdemServico.detentora.ilike(f'%{empresa}%'))

        ordens = ordens.order_by(OrdemServico.data_emissao.desc()).all()

        GRUPOS = {1: 'Capital/RMSP', 2: 'Interior', 3: 'Litoral'}
        STATUS_LABELS_PT = {
            'emitida': 'Emitida', 'enviada_empresa': 'Ag. Empresa',
            'em_revisao': 'Em Revisão', 'aceita': 'Aceita',
            'em_execucao': 'Em Execução', 'executada': 'Executada',
            'recusada': 'Recusada', 'cancelada': 'Cancelada',
        }

        def parse_valor(s):
            try:
                v = str(s or '0').strip()
                if v in ('', '0', '__'):
                    return 0.0
                if ',' in v:
                    # Formato BR "1.234,56": ponto é milhar, vírgula é decimal
                    v = v.replace('.', '').replace(',', '.')
                # Sem vírgula (ex: "1234.56"): ponto já é o separador decimal
                return float(v)
            except (ValueError, TypeError):
                return 0.0

        def custo_cat(itens, slug):
            return round(sum(
                (float(i.quantidade_total or 0)) * parse_valor(i.valor_unitario)
                for i in itens if (i.categoria or '').lower() == slug
            ), 2)

        # ── Workbook ─────────────────────────────────────
        wb = openpyxl.Workbook()

        # ── Aba 1: Resumo por Evento ─────────────────────
        ws = wb.active
        ws.title = 'Eventos'

        COR_H  = 'FF1E3A5F'   # azul escuro
        COR_GR = 'FF2E7D32'   # verde (executada)
        COR_AM = 'FFEF6C00'   # laranja (em_execucao)
        COR_AZ = 'FF1565C0'   # azul (aceita)
        COR_CZ = 'FF546E7A'   # cinza (emitida)

        thin = Side(style='thin', color='FFCCCCCC')
        borda = Border(left=thin, right=thin, top=thin, bottom=thin)
        fmt_moeda = 'R$ #,##0.00'

        # Título
        ws.merge_cells('A1:N1')
        tc = ws['A1']
        tc.value = 'Relatório de Eventos — Módulo Organização'
        tc.font = Font(bold=True, size=14, color='FFFFFFFF')
        tc.fill = PatternFill('solid', fgColor=COR_H)
        tc.alignment = Alignment(horizontal='center', vertical='center')
        ws.row_dimensions[1].height = 30

        ws.merge_cells('A2:N2')
        dc = ws['A2']
        dc.value = f'Gerado em: {datetime.now().strftime("%d/%m/%Y às %H:%M")}'
        dc.font = Font(italic=True, size=10, color='FF6B7280')
        dc.alignment = Alignment(horizontal='center')

        # Totais
        total_pessoas = 0
        custo_geral = 0.0
        dados_planilha = []
        for os in ordens:
            c_mt = custo_cat(os.itens, 'montagem_decoracao')
            c_rh = custo_cat(os.itens, 'recursos_humanos')
            c_ti = custo_cat(os.itens, 'equipamento_informatica')
            c_mg = custo_cat(os.itens, 'material_grafico_expediente')
            c_total = round(c_mt + c_rh + c_ti + c_mg, 2)
            qtd = os.qtd_pessoas_atendidas or 0
            c_pp = round(c_total / qtd, 2) if qtd > 0 else 0.0
            total_pessoas += qtd
            custo_geral += c_total
            dados_planilha.append((os, c_mt, c_rh, c_ti, c_mg, c_total, qtd, c_pp))

        ws.merge_cells('A3:N3')
        rc = ws['A3']
        rc.value = (
            f"Eventos: {len(ordens)}   |   Pessoas Atendidas: {total_pessoas:,}   |   "
            f"Custo Total: R$ {custo_geral:,.2f}   |   "
            f"Custo Médio/Evento: R$ {(custo_geral/len(ordens)):,.2f}" if ordens else "Nenhum evento encontrado"
        )
        rc.font = Font(bold=True, size=10)
        rc.fill = PatternFill('solid', fgColor='FFF0F4FF')
        rc.alignment = Alignment(horizontal='center')
        ws.row_dimensions[3].height = 18

        ws.append([])  # linha 4 vazia

        # Cabeçalhos
        headers = [
            'Nº O.S.', 'Evento', 'Data Evento', 'Local', 'Grupo',
            'Empresa', 'Status', 'Pessoas', 'Montagem/Dec. (R$)',
            'Rec. Humanos (R$)', 'Equip. TI (R$)', 'Mat. Gráfico (R$)',
            'CUSTO TOTAL (R$)', 'Custo/Pessoa (R$)',
        ]
        ws.append(headers)
        hr = ws.max_row
        for ci in range(1, len(headers) + 1):
            cell = ws.cell(row=hr, column=ci)
            cell.font = Font(bold=True, color='FFFFFFFF', size=10)
            cell.fill = PatternFill('solid', fgColor=COR_H)
            cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
            cell.border = borda
        ws.row_dimensions[hr].height = 32

        STATUS_BG = {
            'executada': 'FFE8F5E9', 'em_execucao': 'FFFFF3E0',
            'aceita': 'FFE3F2FD', 'enviada_empresa': 'FFFFF8E1',
            'emitida': 'FFF3F4F6', 'recusada': 'FFFFEBEE', 'cancelada': 'FFF5F5F5',
        }

        for os, c_mt, c_rh, c_ti, c_mg, c_total, qtd, c_pp in dados_planilha:
            bg = STATUS_BG.get(os.status or 'emitida', 'FFFFFFFF')
            row_fill = PatternFill('solid', fgColor=bg)
            row_data = [
                os.numero_os,
                os.evento or '-',
                os.data or '-',
                (os.local or '-')[:60],
                GRUPOS.get(os.regiao_estoque, f'Grupo {os.regiao_estoque}'),
                (os.detentora or '-')[:50],
                STATUS_LABELS_PT.get(os.status, os.status or '-'),
                qtd,
                c_mt, c_rh, c_ti, c_mg, c_total, c_pp,
            ]
            ws.append(row_data)
            ri = ws.max_row
            for ci in range(1, len(headers) + 1):
                cell = ws.cell(row=ri, column=ci)
                cell.fill = row_fill
                cell.border = borda
                cell.alignment = Alignment(vertical='center')
            # Formato moeda para colunas 9-14
            for ci in range(9, 15):
                ws.cell(row=ri, column=ci).number_format = fmt_moeda
            # Destaque custo total
            ct_cell = ws.cell(row=ri, column=13)
            ct_cell.font = Font(bold=True)

        # Larguras
        for ci, w in enumerate([10, 32, 16, 30, 18, 36, 16, 10, 20, 20, 16, 16, 20, 18], 1):
            ws.column_dimensions[get_column_letter(ci)].width = w
        ws.freeze_panes = f'A{hr + 1}'

        # ── Aba 2: Detalhamento por Item ─────────────────
        ws2 = wb.create_sheet('Itens por Evento')
        CAT_LABELS = {
            'montagem_decoracao': 'Montagem/Decoração',
            'recursos_humanos': 'Recursos Humanos',
            'equipamento_informatica': 'Equipamentos TI',
            'material_grafico_expediente': 'Material Gráfico',
        }

        ws2.merge_cells('A1:J1')
        t2 = ws2['A1']
        t2.value = 'Detalhamento de Itens por Evento — Módulo Organização'
        t2.font = Font(bold=True, size=13, color='FFFFFFFF')
        t2.fill = PatternFill('solid', fgColor=COR_H)
        t2.alignment = Alignment(horizontal='center', vertical='center')
        ws2.row_dimensions[1].height = 26

        ws2.append([])

        h2 = ['Nº O.S.', 'Evento', 'Grupo', 'Categoria', 'Descrição do Item',
              'Unid.', 'Diárias', 'Qtd. Total', 'Valor Unit. (R$)', 'Custo Item (R$)']
        ws2.append(h2)
        h2r = ws2.max_row
        for ci in range(1, len(h2) + 1):
            cell = ws2.cell(row=h2r, column=ci)
            cell.font = Font(bold=True, color='FFFFFFFF', size=10)
            cell.fill = PatternFill('solid', fgColor=COR_H)
            cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
            cell.border = borda
        ws2.row_dimensions[h2r].height = 24

        CAT_BG = {
            'montagem_decoracao': 'FFFFF3E0',
            'recursos_humanos': 'FFE3F2FD',
            'equipamento_informatica': 'FFE8F5E9',
            'material_grafico_expediente': 'FFF3E5F5',
        }

        for os, *_ in dados_planilha:
            grp_nome = GRUPOS.get(os.regiao_estoque, f'Grupo {os.regiao_estoque}')
            for item in os.itens:
                val_unit = parse_valor(item.valor_unitario)
                qtd_t = float(item.quantidade_total or 0)
                custo_item = round(val_unit * qtd_t, 2)
                cat_slug = (item.categoria or '').lower()
                bg2 = CAT_BG.get(cat_slug, 'FFFFFFFF')

                ws2.append([
                    os.numero_os,
                    os.evento or '-',
                    grp_nome,
                    CAT_LABELS.get(cat_slug, item.categoria or '-'),
                    item.descricao or '-',
                    item.unidade or '-',
                    item.diarias or 1,
                    qtd_t,
                    val_unit,
                    custo_item,
                ])
                ri2 = ws2.max_row
                row_fill2 = PatternFill('solid', fgColor=bg2)
                for ci in range(1, len(h2) + 1):
                    cell = ws2.cell(row=ri2, column=ci)
                    cell.fill = row_fill2
                    cell.border = borda
                    cell.alignment = Alignment(vertical='center')
                ws2.cell(row=ri2, column=9).number_format = fmt_moeda
                ws2.cell(row=ri2, column=10).number_format = fmt_moeda

        for ci, w in enumerate([10, 28, 18, 22, 45, 10, 10, 12, 18, 18], 1):
            ws2.column_dimensions[get_column_letter(ci)].width = w
        ws2.freeze_panes = f'A{h2r + 1}'

        buffer = BytesIO()
        wb.save(buffer)
        buffer.seek(0)

        return send_file(
            buffer,
            mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
            as_attachment=True,
            download_name=f'relatorio_organizacao_{datetime.now().strftime("%Y%m%d_%H%M%S")}.xlsx'
        )
    except Exception as e:
        return jsonify({'success': False, 'error': str(e)}), 500


@relatorios_bp.route('/api/relatorios/pdf/estoque', methods=['GET'])
@login_requerido
def gerar_pdf_estoque():
    """
    Gera PDF do relatório de posição de estoque
    """
    try:
        modulo = request.args.get('modulo', 'coffee')
        regiao = request.args.get('regiao')
        categoria_id = request.args.get('categoria_id')
        
        # Buscar dados
        query = db.session.query(
            Item.descricao,
            Item.unidade,
            Categoria.nome.label('categoria_nome'),
            EstoqueRegional.regiao_numero,
            EstoqueRegional.quantidade_inicial,
            EstoqueRegional.quantidade_gasto
        ).join(Categoria, Item.categoria_id == Categoria.id)\
         .join(EstoqueRegional, Item.id == EstoqueRegional.item_id)\
         .filter(Categoria.modulo == modulo)
        
        if categoria_id:
            query = query.filter(Item.categoria_id == int(categoria_id))
        if regiao:
            query = query.filter(EstoqueRegional.regiao_numero == int(regiao))
        
        resultados = query.order_by(Categoria.nome, Item.descricao).all()
        
        # Criar PDF
        buffer = BytesIO()
        doc = SimpleDocTemplate(buffer, pagesize=landscape(A4), topMargin=1.5*cm, bottomMargin=1.5*cm)
        elementos = []
        styles = getSampleStyleSheet()
        
        # Título
        titulo_style = ParagraphStyle(
            'CustomTitle',
            parent=styles['Heading1'],
            fontSize=16,
            textColor=colors.HexColor('#1a202c'),
            spaceAfter=20,
            alignment=1  # Centralizado
        )
        
        titulo = f"Relatório de Posição de Estoque"
        if regiao:
            titulo += f" - Região {regiao}"
        elementos.append(Paragraph(titulo, titulo_style))
        elementos.append(Spacer(1, 0.5*cm))
        
        # Data de geração
        data_geracao = datetime.now().strftime('%d/%m/%Y às %H:%M')
        elementos.append(Paragraph(f"Gerado em: {data_geracao}", styles['Normal']))
        elementos.append(Spacer(1, 0.5*cm))
        
        # Tabela
        dados_tabela = [['Categoria', 'Item', 'Unidade', 'Região', 'Inicial', 'Gasto', 'Disponível', '% Uso']]
        
        total_inicial = 0
        total_gasto = 0
        
        for r in resultados:
            # ✅ Tratamento seguro de valores
            try:
                # Converter valores nulos ou inválidos para 0
                inicial_str = str(r.quantidade_inicial or '0').strip()
                gasto_str = str(r.quantidade_gasto or '0').strip()
                
                # Evitar valores inválidos como '__'
                if not inicial_str or inicial_str == '__' or not inicial_str.replace(',', '').replace('.', '').replace('-', ''):
                    inicial = 0
                else:
                    inicial = float(inicial_str.replace('.', '').replace(',', '.'))
                
                if not gasto_str or gasto_str == '__' or not gasto_str.replace(',', '').replace('.', '').replace('-', ''):
                    gasto = 0
                else:
                    gasto = float(gasto_str.replace('.', '').replace(',', '.'))
            except (ValueError, AttributeError):
                # Se ainda houver erro, usar 0
                inicial = 0
                gasto = 0
            
            disponivel = inicial - gasto
            percentual = (gasto / inicial * 100) if inicial > 0 else 0
            
            total_inicial += inicial
            total_gasto += gasto
            
            dados_tabela.append([
                r.categoria_nome,
                r.descricao[:40],  # Limitar tamanho
                r.unidade,
                str(r.regiao_numero),
                f"{inicial:,.0f}",
                f"{gasto:,.0f}",
                f"{disponivel:,.0f}",
                f"{percentual:.1f}%"
            ])
        
        # Linha de total
        total_disponivel = total_inicial - total_gasto
        percentual_total = (total_gasto / total_inicial * 100) if total_inicial > 0 else 0
        dados_tabela.append([
            'TOTAL', '', '', '',
            f"{total_inicial:,.0f}",
            f"{total_gasto:,.0f}",
            f"{total_disponivel:,.0f}",
            f"{percentual_total:.1f}%"
        ])
        
        tabela = Table(dados_tabela, colWidths=[4*cm, 6*cm, 2*cm, 1.5*cm, 2*cm, 2*cm, 2*cm, 2*cm])
        tabela.setStyle(TableStyle([
            ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#667eea')),
            ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
            ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
            ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
            ('FONTSIZE', (0, 0), (-1, 0), 10),
            ('BOTTOMPADDING', (0, 0), (-1, 0), 12),
            ('BACKGROUND', (0, -1), (-1, -1), colors.HexColor('#e2e8f0')),
            ('FONTNAME', (0, -1), (-1, -1), 'Helvetica-Bold'),
            ('GRID', (0, 0), (-1, -1), 1, colors.grey),
            ('FONTSIZE', (0, 1), (-1, -1), 8),
        ]))
        
        elementos.append(tabela)
        doc.build(elementos)
        
        buffer.seek(0)
        return send_file(
            buffer,
            mimetype='application/pdf',
            as_attachment=True,
            download_name=f'relatorio_estoque_{datetime.now().strftime("%Y%m%d_%H%M%S")}.pdf'
        )
        
    except Exception as e:
        return jsonify({'success': False, 'error': str(e)}), 500


@relatorios_bp.route('/api/relatorios/pdf/ordens-servico', methods=['GET'])
@login_requerido
def gerar_pdf_ordens_servico():
    """
    Gera PDF do relatório de Ordens de Serviço
    """
    try:
        modulo = request.args.get('modulo', 'coffee')
        data_inicio = request.args.get('data_inicio')
        data_fim = request.args.get('data_fim')
        regiao = request.args.get('regiao')
        
        query = OrdemServico.query.filter_by(modulo=modulo)
        
        if data_inicio:
            query = query.filter(OrdemServico.data_emissao >= datetime.strptime(data_inicio, '%Y-%m-%d'))
        if data_fim:
            # ✅ Ajustar para o final do dia (23:59:59)
            dt_fim = datetime.strptime(data_fim, '%Y-%m-%d').replace(hour=23, minute=59, second=59, microsecond=999999)
            query = query.filter(OrdemServico.data_emissao <= dt_fim)
        if regiao:
            query = query.filter(OrdemServico.regiao_estoque == int(regiao))
        
        ordens = query.order_by(OrdemServico.data_emissao.desc()).all()
        
        # Criar PDF
        buffer = BytesIO()
        doc = SimpleDocTemplate(buffer, pagesize=landscape(A4), topMargin=1.5*cm, bottomMargin=1.5*cm)
        elementos = []
        styles = getSampleStyleSheet()
        
        # Título
        titulo_style = ParagraphStyle(
            'CustomTitle',
            parent=styles['Heading1'],
            fontSize=16,
            textColor=colors.HexColor('#1a202c'),
            spaceAfter=20,
            alignment=1
        )
        
        titulo = "Relatório de Ordens de Serviço"
        if data_inicio and data_fim:
            titulo += f" ({data_inicio} a {data_fim})"
        elementos.append(Paragraph(titulo, titulo_style))
        elementos.append(Spacer(1, 0.5*cm))
        
        # Resumo
        elementos.append(Paragraph(f"Total de O.S.: {len(ordens)}", styles['Normal']))
        elementos.append(Spacer(1, 0.3*cm))
        
        # Tabela
        dados_tabela = [['Nº O.S.', 'Data Emissão', 'Serviço', 'Evento', 'Contratada', 'Região', 'Itens']]
        
        for os in ordens:
            total_itens = len(os.itens)
            data_emissao = os.data_emissao.strftime('%d/%m/%Y') if os.data_emissao else '-'
            
            dados_tabela.append([
                os.numero_os,
                data_emissao,
                os.servico[:25] if os.servico else '-',
                os.evento[:25] if os.evento else '-',
                os.detentora[:25] if os.detentora else '-',
                str(os.regiao_estoque) if os.regiao_estoque else '-',
                str(total_itens)
            ])
        
        tabela = Table(dados_tabela, colWidths=[2.5*cm, 2.5*cm, 4*cm, 4*cm, 4*cm, 1.5*cm, 1.5*cm])
        tabela.setStyle(TableStyle([
            ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#667eea')),
            ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
            ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
            ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
            ('FONTSIZE', (0, 0), (-1, 0), 10),
            ('BOTTOMPADDING', (0, 0), (-1, 0), 12),
            ('GRID', (0, 0), (-1, -1), 1, colors.grey),
            ('FONTSIZE', (0, 1), (-1, -1), 8),
        ]))
        
        elementos.append(tabela)
        doc.build(elementos)
        
        buffer.seek(0)
        return send_file(
            buffer,
            mimetype='application/pdf',
            as_attachment=True,
            download_name=f'relatorio_os_{datetime.now().strftime("%Y%m%d_%H%M%S")}.pdf'
        )
        
    except Exception as e:
        return jsonify({'success': False, 'error': str(e)}), 500
